"""Convert CSV -> Excel (XLSX) with AI-enhanced formatting.

Modes:
- 'basic': Simple conversion (default, fast, no AI)
- 'ai': Intelligent formatting with column type detection, styling, and suggestions
"""
import pandas as pd
import logging
from pathlib import Path
from typing import Optional, Dict, Any

logger = logging.getLogger(__name__)


def convert(input_path: str, output_path: str, mode: str = 'basic') -> None:
    """Convert CSV to Excel with optional AI enhancements.
    
    Args:
        input_path: Path to input CSV
        output_path: Path to output XLSX
        mode: Conversion mode ('basic' or 'ai')
    """
    if mode == 'basic':
        _convert_basic(input_path, output_path)
    elif mode == 'ai':
        _convert_with_ai(input_path, output_path)
    else:
        raise ValueError(f"Unknown mode: {mode}. Use 'basic' or 'ai'.")


def _convert_basic(input_path: str, output_path: str) -> None:
    """Basic CSV to Excel conversion without AI.
    
    Args:
        input_path: Path to input CSV
        output_path: Path to output XLSX
    """
    logger.info(f"Converting CSV to Excel (basic mode)...")
    df = pd.read_csv(input_path)
    df.to_excel(output_path, index=False)
    logger.info(f"✓ Conversion complete: {output_path}")


def _convert_with_ai(input_path: str, output_path: str) -> None:
    """AI-enhanced CSV to Excel conversion with intelligent formatting.
    
    Args:
        input_path: Path to input CSV
        output_path: Path to output XLSX
    """
    logger.info("🤖 Starting AI-enhanced CSV to Excel conversion...")
    
    try:
        from utils.groq_service import GroqDocumentReconstructor
        from utils.ai_prompts import CSV_TO_EXCEL_ENHANCEMENT_SYSTEM, format_prompt
        from utils.privacy import PrivacyChecker
    except ImportError as e:
        logger.warning(f"AI mode unavailable: {e}")
        logger.info("Falling back to basic mode...")
        return _convert_basic(input_path, output_path)
    
    # Phase 1: Load CSV
    logger.info("📊 Phase 1: Loading CSV data...")
    df = pd.read_csv(input_path)
    
    # Phase 2: Privacy check
    logger.info("🔒 Phase 2: Privacy check...")
    checker = PrivacyChecker()
    sample_text = df.head(20).to_string()
    has_sensitive, findings = checker.check_text(sample_text)
    
    if has_sensitive:
        logger.warning("⚠️  Sensitive data detected in CSV:")
        for finding in findings:
            logger.warning(f"   - {finding}")
        logger.warning("Consider using basic mode for sensitive data.")
        
        import sys
        if sys.stdin.isatty():
            response = input("Continue with AI formatting? (y/N): ").strip().lower()
            if response != 'y':
                logger.info("Using basic mode instead...")
                return _convert_basic(input_path, output_path)
    
    # Phase 3: AI Analysis
    logger.info("🧠 Phase 3: AI formatting analysis...")
    
    # Prepare data for AI
    csv_preview = df.head(10).to_dict(orient='records')
    column_info = {
        col: {
            'dtype': str(df[col].dtype),
            'null_count': int(df[col].isnull().sum()),
            'unique_count': int(df[col].nunique()),
            'sample_values': df[col].dropna().head(5).tolist()
        }
        for col in df.columns
    }
    
    try:
        groq = GroqDocumentReconstructor()
        
        prompt = format_prompt(
            """Analyze this CSV data for intelligent Excel formatting.

CSV PREVIEW:
{csv_preview}

COLUMN INFO:
{column_info}

Provide JSON with:
1. column_formats: {{column_name: {{type: "text|number|date|currency", format: "format_string"}}}}
2. header_style: {{bold: true, bg_color: "#color"}}
3. conditional_formats: [{{column: "col", rule: "rule", format: "format"}}]
4. suggested_charts: [{{type: "chart_type", columns: ["col1", "col2"]}}]

Be practical and professional.""",
            csv_preview=str(csv_preview)[:1000],
            column_info=str(column_info)[:1000]
        )
        
        response = groq.client.chat.completions.create(
            model=groq.config.model,
            messages=[
                {"role": "system", "content": CSV_TO_EXCEL_ENHANCEMENT_SYSTEM},
                {"role": "user", "content": prompt}
            ],
            temperature=0.1,
            max_tokens=2000
        )
        
        ai_suggestions = response.choices[0].message.content or "{}"
        logger.info("✓ AI analysis complete")
        
    except Exception as e:
        logger.warning(f"AI analysis failed: {e}, using basic formatting")
        ai_suggestions = "{}"
    
    # Phase 4: Apply formatting
    logger.info("📝 Phase 4: Applying Excel formatting...")
    _apply_enhanced_formatting(df, output_path, ai_suggestions)
    
    logger.info(f"✓ AI-enhanced conversion complete: {output_path}")


def _apply_enhanced_formatting(
    df: pd.DataFrame,
    output_path: str,
    ai_suggestions: str
) -> None:
    """Apply AI-suggested formatting to Excel output.
    
    Args:
        df: DataFrame to export
        output_path: Output Excel path
        ai_suggestions: JSON string with AI suggestions
    """
    import json
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    
    # First, save basic Excel
    df.to_excel(output_path, index=False, engine='openpyxl')
    
    # Parse AI suggestions
    try:
        suggestions = json.loads(ai_suggestions.strip().replace('```json', '').replace('```', ''))
    except:
        logger.warning("Could not parse AI suggestions, using default formatting")
        suggestions = {}
    
    # Load workbook for styling
    wb = load_workbook(output_path)
    ws = wb.active
    
    if ws is None:
        logger.warning("Could not access worksheet for styling")
        return
    
    # Apply header styling
    header_style = suggestions.get('header_style', {})
    if header_style.get('bold', True):
        for cell in ws[1]:
            cell.font = Font(bold=True, size=11)
            cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter  # type: ignore
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    wb.save(output_path)
    logger.info("✓ Enhanced formatting applied")


if __name__ == '__main__':
    import sys
    if len(sys.argv) < 3:
        print('Usage: python csv_to_excel.py input.csv output.xlsx [mode]')
        print('Modes: basic (default), ai')
        sys.exit(2)
    
    inp = sys.argv[1]
    out = sys.argv[2]
    mode = sys.argv[3] if len(sys.argv) > 3 else 'basic'
    
    logging.basicConfig(level=logging.INFO, format='[%(levelname)s] %(message)s')
    convert(inp, out, mode=mode)

